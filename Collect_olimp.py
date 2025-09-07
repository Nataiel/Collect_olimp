import tkinter as tk
from tkinter import filedialog, messagebox, ttk
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Border, Side, Alignment
import os
from datetime import datetime

path_file = None  # переменная для хранения пути к файлу
olimp_subjects = ['Выберите предмет']  # переменная для списка предметов из файла
sheet_names = []  # Список листов в Исходном файле
class_entry = []  # Перечень Entry для создания параллелей [0:4, 1:5, 2:6, 3:7, 4:8, 5:9, 6:10, 7:11]


def open_file():
    """Функция для открытия файла и сохранения пути"""
    global path_file, olimp_subjects, sheet_names

    # Запрашиваем файл у пользователя
    file_path = filedialog.askopenfilename(title="Выберите Excel файл",
                                           filetypes=[("Excel files", "*.xlsx"), ("All files", "*.*")])

    # Если пользователь выбрал файл (не нажал "Отмена")
    if file_path:
        path_file = file_path
        # Обновляем текст в метке
        file_label.config(text=f"Выбран файл: {os.path.basename(path_file)}")
        status_label.config(text="Файл успешно загружен", fg="green")
        # Активируем кнопку модификации
        modify_button.config(state=tk.NORMAL)
        print(f"Файл выбран: {path_file}")  # Для отладки

        wb = load_workbook(path_file)
        sheet_names = wb.sheetnames
        sheet_names = [x for x in sheet_names if any(y.isdigit() for y in x) and 'Лист' not in x]

        for x in sheet_names:
            collect_subjects(wb[x])
        olimp_subjects = [x for x in olimp_subjects if not ('умма' in x or x == 'Выберите предмет')]
        subject_var.config(values=olimp_subjects)


def modify_file():
    """Функция для модификации файла"""
    global path_file, olimp_subjects, sheet_names

    try:
        if subject_var.get() == 'Выберите предмет':
            return

        # Загружаем workbook
        wb = load_workbook(path_file)

        if subject_var.get() in wb.sheetnames:
            del wb[subject_var.get()]

        wb.create_sheet(subject_var.get())
        sheet = wb[subject_var.get()]

        sheet.cell(row=1, column=1).value = '№'
        sheet.cell(row=1, column=2).value = 'Фамилия'
        sheet.cell(row=1, column=3).value = 'Имя'
        sheet.cell(row=1, column=4).value = 'Отчество'
        sheet.cell(row=1, column=5).value = 'Класс'
        row_pos = 2

        for y in sheet_names:
            col = 6
            row = 1
            for x in range(col, 100):
                if wb[y].cell(row=1, column=x).value == subject_var.get():
                    break

            row += 1

            for z in range(row, 100):
                if wb[y].cell(row=z, column=x).value:
                    wb[subject_var.get()].cell(row=row_pos, column=1).value = row_pos - 1
                    wb[subject_var.get()].cell(row=row_pos, column=2).value = wb[y].cell(row=z, column=2).value
                    wb[subject_var.get()].cell(row=row_pos, column=3).value = wb[y].cell(row=z, column=3).value
                    wb[subject_var.get()].cell(row=row_pos, column=4).value = wb[y].cell(row=z, column=4).value
                    wb[subject_var.get()].cell(row=row_pos, column=5).value = y
                    row_pos += 1

        # Сохраняем изменения
        wb.save(path_file)

        messagebox.showinfo("Успех", "Файл успешно модифицирован!")
        status_label.config(text="Файл модифицирован", fg="blue")

    except Exception as e:
        messagebox.showerror("Ошибка", f"Не удалось модифицировать файл:\n{str(e)}")
        status_label.config(text="Ошибка модификации", fg="red")


# Функция сбора предметов с листа
def collect_subjects(sheet):
    global olimp_subjects

    col = 6
    row = 1
    for x in range(col, 100):
        aux = sheet.cell(row=row, column=x).value
        if aux:
            if aux not in olimp_subjects:
                olimp_subjects.append(aux)


def close_app():
    """Функция для закрытия приложения"""
    if messagebox.askokcancel("Выход", "Вы уверены, что хотите выйти?"):
        root.destroy()


def create_file():
    global class_entry

    alphabet = 'АБВГДЕЖЗИКЛМНОПРСТУФХЦЧШЩЫЭЮЯ'
    create_file_path = filedialog.asksaveasfilename(title="Сохранить файл как", defaultextension=".xlsx",
                                                    filetypes=[("Excel files", "*.xlsx"), ("All files", "*.*")],
                                                    initialfile=f"Выбор ВсОШ ШЭ {datetime.now().year}.xlsx")

    wb = Workbook()
    class_number = 3
    for x in class_entry:
        class_number += 1
        if not x.get():
            continue

        for y in range(int(x.get())):
            wb.create_sheet(f'{class_number}{alphabet[y]}')

    class_sheet_names = wb.sheetnames
    if 'Sheet' in class_sheet_names and len(class_sheet_names) > 1:
        del wb['Sheet']

    class_sheet_names = wb.sheetnames

    wb = class_table_creater(wb, class_sheet_names)

    wb.save(create_file_path)


def class_table_creater(wb, sheets):
    olimp = ["русский язык", "математика", "астрономия", "биология", "география", "информатика", "искусство (МХК)",
             "история", "литература", "обществознание", "ОБ ЗР", "право", "труд (технология)", "физика",
             "физическая культура", "химия", "экология", "экономика", "английский язык", "испанский язык",
             "итальянский язык", "китайский язык", "немецкий язык", "французский язык"]

    columns_width = {
        'A': 3,  # № п/п
        'B': 30,  # Фамилия
        'C': 20,  # Имя
        'D': 20,  # Отчество
        'E': 10,  # Дом/Школа
        'F': 5,  # Предмет 1
        'G': 5,  # Предмет 2
        'H': 5,  # Предмет 3
        'I': 5,  # Предмет 4
        'J': 5,  # Предмет 5
        'K': 5,  # Предмет 6
        'L': 5,  # Предмет 7
        'M': 5,  # Предмет 8
        'N': 5,  # Предмет 9
        'O': 5,  # Предмет 10
        'P': 5,  # Предмет 11
        'Q': 5,  # Предмет 12
        'R': 5,  # Предмет 13
        'S': 5,  # Предмет 14
        'T': 5,  # Предмет 15
        'U': 5,  # Предмет 16
        'V': 5,  # Предмет 17
        'W': 5,  # Предмет 18
        'X': 5,  # Предмет 19
        'Y': 5,  # Предмет 20
        'Z': 5,  # Предмет 21
        'AA': 5,  # Предмет 22
        'AB': 5,  # Предмет 23
        'AC': 5  # Предмет 24
    }

    # Настройка стилей
    border_style = Side(border_style="thin", color="000000")
    border = Border(left=border_style, right=border_style, top=border_style, bottom=border_style)
    alignment_center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    alignment_left = Alignment(horizontal="left", vertical="center", wrap_text=True, textRotation=90)

    for x in sheets:
        ws = wb[x]

        for col, width in columns_width.items():
            ws.column_dimensions[col].width = width
        ws.row_dimensions[1].height = 80

        if '4' in x:
            olimps = olimp[:2]
        else:
            olimps = olimp[:]

        row = 1
        col = 1
        for y in ['№', 'Фамилия', 'Имя', 'Отчество', 'Дома/Школа'] + olimps:
            ws.cell(row=row, column=col).value = y
            ws.cell(row=row, column=col).border = border
            if col > 5:
                ws.cell(row=row, column=col).alignment = alignment_left
            else:
                ws.cell(row=row, column=col).alignment = alignment_center
            col += 1

    return wb


# Создаем главное окно
root = tk.Tk()
root.title("Обработка назначений на ВсОШ ШЭ ©")
root.geometry("500x300")
root.resizable(False, False)

left = tk.LabelFrame(root, text='Количество классов')
right = tk.LabelFrame(root, text='Обработка файла')

# Исходные данные для файла Образца

ss = 2
ww = 5
for x in ['4', '5', '6', '7', '8', '9', '10', '11']:
    aux_frame = tk.Frame(left)
    tk.Label(aux_frame, text=x, font=('Arial', 10)).pack(side=tk.LEFT, padx=5, anchor='w')
    class_entry.append(tk.Entry(aux_frame, width=ww, font=('Arial', 10)))
    class_entry[-1].pack(side=tk.LEFT, padx=5, anchor='w', pady=ss)
    aux_frame.pack(side=tk.TOP)

# Кнопка создания файла образца
button_create = tk.Button(left, text='Создать образец', command=create_file, font=('Arial', 10), width=15)
button_create.pack(anchor='n', pady=5, padx=5)

# Стилизация
button_style = {'font': ('Arial', 11), 'width': 13, 'height': 1, 'bg': '#f0f0f0'}

# Создаем фрейм для кнопок
button_frame = tk.Frame(right)
button_frame.pack(pady=5, padx=5)

# Кнопка открытия файла
open_button = tk.Button(button_frame, text="Открыть файл", command=open_file, **button_style)
open_button.pack(pady=5, padx=5)

# Фрейм для выпадающего списка и кнопки модификации
modify_frame = tk.Frame(button_frame)
modify_frame.pack(pady=5, padx=5)

# Выпадающий список слева
subject_var = ttk.Combobox(modify_frame, values=olimp_subjects, state="readonly", height=15)
subject_var.set(olimp_subjects[0])  # значение по умолчанию
subject_var.config(font=('Arial', 11), width=17, height=15)
subject_var.pack(side=tk.LEFT, padx=5)

# Кнопка модификации (изначально неактивна)
modify_button = tk.Button(modify_frame, text="Модифицировать", command=modify_file, state=tk.DISABLED, **button_style)
modify_button.pack(side=tk.LEFT)

# Кнопка закрытия
close_button = tk.Button(button_frame, text="Закрыть", command=close_app, **button_style)
close_button.pack(pady=5, padx=5)

# Метка для отображения выбранного файла
file_label = tk.Label(right, text="Файл не выбран", font=('Arial', 10), fg='gray')
file_label.pack(pady=5, padx=5)

# Метка статуса
status_label = tk.Label(right, text="Готов к работе", font=('Arial', 10), fg='green')
status_label.pack(pady=5, padx=5)

left.pack(side=tk.LEFT, pady=5, padx=5, anchor='n')
right.pack(pady=5, padx=5, anchor='n', fill='x')

# Запускаем главный цикл
root.mainloop()
